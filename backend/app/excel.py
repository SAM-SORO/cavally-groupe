"""Generation du devis .xlsx (openpyxl).

La *mise en forme* est fixe — c'est l'identite du devis Cavally Livres, calee sur
`un_exemple.xlsx` : en-tete etablissement, bloc d'identification, tableau
Categorie | Designation | Qte | Prix Unitaire | Total HT, ligne de total, puis
consignes.

Le *contenu*, lui, vient integralement du document uploade : les rubriques sont
celles du document, les consignes sont celles du document (aucune valeur par
defaut), et chaque classe traitee donne son propre onglet. Une information
absente reste une case a completer, jamais une valeur inventee.

Deux regles non negociables :
- la colonne "Prix Unitaire" est laissee VIDE (remplie a la main dans le fichier) ;
- "Total HT" et "TOTAL ESTIMATIF HT" sont de vraies formules Excel, jamais des
  valeurs figees, pour que tout se recalcule des qu'un prix est saisi.
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .schemas import Extraction, ListeClasse

# --------------------------------------------------------------------------- #
# Charte Cavally Livres (code couleur.txt) — aucune couleur inventee.
# --------------------------------------------------------------------------- #

JAUNE = "FFFFB800"  # #FFB800
NOIR = "FF000000"  # #000000
BLANC = "FFFFFFFF"  # #FFFFFF

# Neutres = noir desature, et jaune de la charte pose en aplat transparent
# sur du blanc. Aucune teinte hors charte n'est introduite.
GRIS_TEXTE = "FF555555"  # noir a 67% sur blanc
GRIS_LEGER = "FF666666"  # noir a 60% sur blanc
GRIS_BORDURE = "FFE6E6E6"  # noir a 10% sur blanc
GRIS_BANDE = "FFFAFAFA"  # noir a 2% sur blanc
JAUNE_VOILE = "FFFFF4D9"  # #FFB800 a 15% sur blanc

FONT = "Calibri"
FORMAT_FCFA = '#,##0\\ "FCFA"'
FORMAT_QTE = "#,##0"

A_COMPLETER = "_______________________"

PREMIERE_LIGNE_TABLEAU = 9  # ligne des en-tetes, comme dans le fichier d'exemple
COLONNES = [
    ("Catégorie", 16, "center"),
    ("Désignation des Articles / Fournitures", 65, "left"),
    ("Qté", 8, "center"),
    ("Prix Unitaire (FCFA)", 20, "right"),
    ("Total HT (FCFA)", 22, "right"),
]

MENTION_EMETTRICE = "Devis estimatif établi par Cavally Livres — les prix unitaires restent à compléter."

_BORDURE = Side(style="thin", color=GRIS_BORDURE)
CADRE = Border(left=_BORDURE, right=_BORDURE, top=_BORDURE, bottom=_BORDURE)
CADRE_TOTAL = Border(
    top=Side(style="thin", color=GRIS_BORDURE),
    bottom=Side(style="double", color=JAUNE),
)

_INTERDITS_ONGLET = r'[\\/*?:\[\]]'


def _f(size: int = 11, *, bold: bool = False, italic: bool = False, color: str = NOIR) -> Font:
    return Font(name=FONT, size=size, bold=bold, italic=italic, color=color)


def _slug(value: str) -> str:
    """Fragment de nom de fichier sur (ASCII, sans separateur de chemin)."""
    ascii_value = "".join(
        c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c)
    )
    ascii_value = ascii_value.encode("ascii", "ignore").decode("ascii")
    ascii_value = re.sub(r"[^A-Za-z0-9]+", "-", ascii_value).strip("-")
    return re.sub(r"-{2,}", "-", ascii_value)[:48]


def build_filename(extraction: Extraction, today: date | None = None) -> str:
    """Nom de fichier lisible, adapte au nombre de classes du document."""
    today = today or date.today()
    morceaux = ["Devis"]

    etablissement = _slug(extraction.etablissement)
    if etablissement:
        morceaux.append(etablissement)

    classes = extraction.classes
    if len(classes) == 1:
        classe = _slug(classes[0])
        if classe:
            morceaux.append(classe)
    elif len(classes) > 1:
        morceaux.append(f"{len(classes)}-classes")

    if len(morceaux) == 1:
        morceaux.append("Fournitures")
    morceaux.append(today.strftime("%Y-%m-%d"))
    return "-".join(morceaux) + ".xlsx"


def _nom_onglet(liste: ListeClasse, index: int, pris: set[str]) -> str:
    """Nom d'onglet valide, unique et lisible (Excel : 31 caracteres max)."""
    base = re.sub(_INTERDITS_ONGLET, " ", liste.classe).strip()
    base = re.sub(r"\s+", " ", base) or (f"Classe {index + 1}" if index else "Devis fournitures")
    nom = base[:31]
    if nom in pris:
        suffixe = 2
        while f"{base[:28]} ({suffixe})" in pris:
            suffixe += 1
        nom = f"{base[:28]} ({suffixe})"
    pris.add(nom)
    return nom


# --------------------------------------------------------------------------- #
# Blocs d'un onglet
# --------------------------------------------------------------------------- #


def _ecrire_entete(ws: Worksheet, extraction: Extraction, liste: ListeClasse, today: date) -> None:
    # Une information absente devient une case a completer, pas une invention.
    ws["A1"] = extraction.etablissement or A_COMPLETER
    ws["A1"].font = _f(14, bold=True, color=NOIR if extraction.etablissement else GRIS_TEXTE)
    ws.row_dimensions[1].height = 22

    ws["E1"] = "DEVIS ESTIMATIF"
    ws["E1"].font = _f(16, bold=True)
    ws["E1"].alignment = Alignment(horizontal="right", vertical="center")

    if extraction.coordonnees:
        ws["A2"] = extraction.coordonnees
        ws["A2"].font = _f(10, color=GRIS_TEXTE)

    for ligne, (libelle, valeur) in enumerate(
        (
            ("Année Scolaire :", extraction.annee_scolaire),
            ("Classe :", liste.classe),
            ("Date :", today.strftime("%d/%m/%Y")),
        ),
        start=3,
    ):
        cellule_libelle = ws.cell(row=ligne, column=5, value=libelle)
        cellule_libelle.font = _f(bold=True)
        cellule_libelle.alignment = Alignment(horizontal="right", vertical="center")
        cellule_valeur = ws.cell(row=ligne, column=6, value=valeur or A_COMPLETER)
        cellule_valeur.font = _f(color=NOIR if valeur else GRIS_TEXTE)

    for ligne, libelle in ((5, "Nom de l'élève :"), (6, "Parent / Tuteur :")):
        ws.cell(row=ligne, column=1, value=libelle).font = _f(bold=True)
        ws.cell(row=ligne, column=2, value=A_COMPLETER).font = _f(color=GRIS_TEXTE)


def _ecrire_tableau(ws: Worksheet, liste: ListeClasse) -> tuple[int, int]:
    """Ecrit l'en-tete du tableau puis les lignes. Renvoie (premiere, derniere)."""
    ligne_entete = PREMIERE_LIGNE_TABLEAU
    ws.row_dimensions[ligne_entete].height = 26

    for index, (titre, largeur, alignement) in enumerate(COLONNES, start=1):
        lettre = get_column_letter(index)
        ws.column_dimensions[lettre].width = largeur
        cellule = ws.cell(row=ligne_entete, column=index, value=titre)
        cellule.font = _f(bold=True)  # noir sur jaune : contraste conforme
        cellule.fill = PatternFill("solid", fgColor=JAUNE)
        cellule.alignment = Alignment(horizontal=alignement, vertical="center", wrap_text=True)
    ws.column_dimensions["F"].width = 18

    premiere = ligne_entete + 1
    for decalage, article in enumerate(liste.articles):
        ligne = premiere + decalage
        bande = decalage % 2 == 1
        remplissage = PatternFill("solid", fgColor=GRIS_BANDE) if bande else None

        valeurs = (
            (1, article.categorie, "center", None),
            (2, article.designation, "left", None),
            (3, article.qte, "center", FORMAT_QTE),
            (4, None, "right", FORMAT_FCFA),  # Prix unitaire : VIDE, saisi a la main
            (5, f"=C{ligne}*D{ligne}", "right", FORMAT_FCFA),  # vraie formule
        )
        for colonne, valeur, alignement, format_nombre in valeurs:
            cellule = ws.cell(row=ligne, column=colonne, value=valeur)
            cellule.font = _f()
            cellule.border = CADRE
            cellule.alignment = Alignment(
                horizontal=alignement, vertical="center", wrap_text=colonne == 1
            )
            if format_nombre:
                cellule.number_format = format_nombre
            if remplissage:
                cellule.fill = remplissage

    return premiere, premiere + len(liste.articles) - 1


def _ecrire_total(ws: Worksheet, premiere: int, derniere: int) -> int:
    ligne = derniere + 2
    ws.row_dimensions[ligne].height = 20

    libelle = ws.cell(row=ligne, column=4, value="TOTAL ESTIMATIF HT :")
    libelle.font = _f(12, bold=True)
    libelle.alignment = Alignment(horizontal="right", vertical="center")
    libelle.fill = PatternFill("solid", fgColor=JAUNE_VOILE)
    libelle.border = CADRE_TOTAL

    total = ws.cell(row=ligne, column=5, value=f"=SUM(E{premiere}:E{derniere})")
    total.font = _f(12, bold=True)
    total.alignment = Alignment(horizontal="right", vertical="center")
    total.number_format = FORMAT_FCFA
    total.fill = PatternFill("solid", fgColor=JAUNE_VOILE)
    total.border = CADRE_TOTAL
    return ligne


def _ecrire_pied(ws: Worksheet, liste: ListeClasse, ligne_total: int) -> None:
    """Consignes du document uniquement — aucune consigne par defaut."""
    ligne = ligne_total + 3

    if liste.consignes:
        titre = ws.cell(row=ligne, column=1, value="CONSIGNES & REMARQUES :")
        titre.font = _f(bold=True)
        for decalage, consigne in enumerate(liste.consignes, start=1):
            cellule = ws.cell(row=ligne + decalage, column=1, value=f"• {consigne}")
            cellule.font = _f(10, italic=True, color=GRIS_LEGER)
        ligne += len(liste.consignes) + 2

    mention = ws.cell(row=ligne, column=1, value=MENTION_EMETTRICE)
    mention.font = _f(9, italic=True, color=GRIS_LEGER)


def _remplir_onglet(ws: Worksheet, extraction: Extraction, liste: ListeClasse, today: date) -> None:
    ws.sheet_view.showGridLines = False

    _ecrire_entete(ws, extraction, liste, today)
    premiere, derniere = _ecrire_tableau(ws, liste)
    ligne_total = _ecrire_total(ws, premiere, derniere)
    _ecrire_pied(ws, liste, ligne_total)

    ws.freeze_panes = f"A{premiere}"
    ws.print_title_rows = f"{PREMIERE_LIGNE_TABLEAU}:{PREMIERE_LIGNE_TABLEAU}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# --------------------------------------------------------------------------- #
# Entree publique
# --------------------------------------------------------------------------- #


def build_workbook(extraction: Extraction, today: date | None = None) -> bytes:
    """Construit le classeur — un onglet par classe — et le renvoie en octets."""
    listes = [liste for liste in extraction.listes if liste.articles]
    if not listes:
        raise ValueError("Aucun article a ecrire dans le devis.")

    today = today or date.today()
    workbook = Workbook()
    workbook.remove(workbook.active)

    pris: set[str] = set()
    for index, liste in enumerate(listes):
        ws = workbook.create_sheet(_nom_onglet(liste, index, pris))
        _remplir_onglet(ws, extraction, liste, today)

    classes = extraction.classes
    workbook.properties.title = "Devis estimatif — fournitures scolaires"
    workbook.properties.creator = "Cavally Livres"
    workbook.properties.description = (
        f"{len(extraction.articles)} article(s) extraits automatiquement"
        + (f" sur {len(classes)} classe(s)" if len(classes) > 1 else "")
        + ". Colonne Prix Unitaire a completer manuellement."
    )

    flux = io.BytesIO()
    workbook.save(flux)
    return flux.getvalue()
